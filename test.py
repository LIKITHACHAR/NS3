import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Border ,Side ,Font

# Create a new workbook
wb = openpyxl.Workbook()

worksheet = wb.active
# Set thick borders for the merged cell
border = Border(left=Side(style='thin'), 
                top=Side(style='thin'), 
                right=Side(style='thin'), 
                bottom=Side(style='thin'))
font = Font(bold=True)

# Define the data and merged cell range
data = ["Shareholder:", "Partners Group Access PF 5556 (UK) L.P."]
data2 = ["Legal form:", "Limited partnership registered with the Registrar of Companies for Scotland under number SL36397"]
data3 = ["Address:","50 Lothian Road, Festival Square, Edinburgh, Scotland, United Kingdom"]
data4 = []
data5 = ["Date","Subscription/acquisition","Shares transferred/ converted/ subscribed","Total number of shares:","Signature"]
merged_range = "B1:E1"
merged_range1 = "B2:E2"
merged_range2 = "B3:E3"
merged_range3 = "A4:E4"


# Change height of row A1
worksheet.row_dimensions[1].height = 30 
worksheet.row_dimensions[2].height = 30
worksheet.row_dimensions[3].height = 30
worksheet.row_dimensions[4].height = 15
worksheet.row_dimensions[5].height = 35
# Change width of column B
worksheet.column_dimensions["A"].width = 16
worksheet.column_dimensions["B"].width = 50
worksheet.column_dimensions["C"].width = 30
worksheet.column_dimensions["D"].width = 15
worksheet.column_dimensions["E"].width = 25


# Write data to cells
worksheet.append(data)
worksheet.append(data2)
worksheet.append(data3)
worksheet.append(data4)
worksheet.append(data5)
# Merge cells and set alignment
worksheet.merge_cells(merged_range)
worksheet.merge_cells(merged_range1)
worksheet.merge_cells(merged_range2)
worksheet.merge_cells(merged_range3)

worksheet['B1'].alignment = Alignment(horizontal ='center', vertical='center')
worksheet['B2'].alignment = Alignment(horizontal ='center', vertical='center')
worksheet['B3'].alignment = Alignment(horizontal ='center', vertical='center')
worksheet['C5'].alignment = Alignment(horizontal ='center', vertical='center',wrap_text=True)
worksheet['D5'].alignment = Alignment(horizontal ='center', vertical='center',wrap_text=True)
worksheet['E5'].alignment = Alignment(horizontal ='center', vertical='center')
worksheet['A5'].alignment = Alignment(horizontal ='center', vertical='center')
worksheet['B5'].alignment = Alignment(horizontal ='center', vertical='center')

#font
worksheet['A1'].font = font
worksheet['A2'].font = font
worksheet['A3'].font = font
worksheet['A5'].font = font
worksheet['B5'].font = font
worksheet['B1'].font = font
worksheet['C5'].font = font
worksheet['D5'].font = font
worksheet['E5'].font = font
#border

worksheet['A5'].border = border
worksheet['B5'].border = border
worksheet['C5'].border = border
worksheet['D5'].border = border
worksheet['E5'].border = border

last_row = worksheet.max_row
print(last_row)

# Save the workbook
wb.save("merged_and_centered.xlsx")

print("Excel file created successfully!")



#import sys
#print ("test")
#print(sys.version)
#print(sys.executable)