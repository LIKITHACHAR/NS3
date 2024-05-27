from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
 
 
 
wb = Workbook()
ws = wb.active
 
data = [
    ['Text1', 'Text2', 'Text3', 'Text4'],
    ['Text1', 'Text2', 'Text3', 'Text4'],
    ['Text1', 'Text2', 'Text3', 'Text4'],
    ['Text1', 'Text2', 'Text3', 'Text4'],
]
 
x = len(data)
 
# add column headings. NB. these must be strings
ws.append(["Header1", "Header2", "Header3", "Header4", "Header5"])
 
tab = Table(displayName="Table1", ref="A1:E5")
 
# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style
 
'''
Table must be added using ws.add_table() method to avoid duplicate names.
Using this method ensures table name is unque through out defined names and all other table name.
'''
ws.add_table(tab)
 
rwstart = 2
rwend = 0
 
for d in data:
    rwend = rwstart + 10
    for i in range(1,len(d)+1):
        ws.merge_cells(start_row=rwstart, start_column=i, end_row=rwend, end_column=i)
        ws.cell(row=rwstart, column=i).value = d[i-1]
        ws.cell(row=rwstart, column=i).alignment = Alignment(horizontal='left',vertical='center')
        # ws.cell(row=rwstart, column=i).alignment = Alignment(vertical='middle')
    rwstart = rwend + 1
       
 
 
wb.save("table.xlsx")

#code logic for last column 

'''rwstart = 2
rwend = 0
 
for d in data:
    rwend = rwstart + 10
    for i in range(1,len(d)+2):
        ws.merge_cells(start_row=rwstart, start_column=i, end_row=rwend, end_column=i)
        if len(d)+2 == 5:  ws.cell(row=rwstart, column=i).value = 'Constant Text'
        else: ws.cell(row=rwstart, column=i).value = d[i-1]
        ws.cell(row=rwstart, column=i).alignment = Alignment(horizontal='left',vertical='center')
    rwstart = rwend + 1'''